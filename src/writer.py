from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def atualizar_planilha_por_data(dados: list[dict], caminho_saida: str):
    """
    Atualiza uma planilha Excel organizando os dados por data.

    Cada aba representa uma data e contém:
    - Coluna A: rótulos "Nome", "Atividades", "Atos"
    - Colunas B em diante: nomes, atividades e atos por pessoa
    """
    try:
        # Abrir planilha existente ou criar nova
        try:
            wb = load_workbook(caminho_saida)
        except FileNotFoundError:
            wb = Workbook()
            wb.remove(wb.active)

        # Estilos
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        header_fill = PatternFill("solid", fgColor="DDDDDD")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # Agrupar dados por data
        datas_dict = {}
        for registro in dados:
            data = registro['Data']
            if data not in datas_dict:
                datas_dict[data] = []
            datas_dict[data].append(registro)

        for data, registros in datas_dict.items():
            nome_aba = data.replace('/', '_').replace('-', '_')[:31]

            if nome_aba not in wb.sheetnames:
                ws = wb.create_sheet(title=nome_aba)
            else:
                ws = wb[nome_aba]

            # Cabeçalhos fixos na primeira coluna
            ws.cell(row=1, column=1, value="Nome").font = bold_font
            ws.cell(row=2, column=1, value="Atividades").font = bold_font
            ws.cell(row=3, column=1, value="Atos").font = bold_font

            for row in range(1, 4):
                cell = ws.cell(row=row, column=1)
                cell.alignment = center_align
                cell.fill = header_fill
                cell.border = thin_border

            # Preencher dados a partir da coluna 2
            for col, pessoa in enumerate(registros, start=2):
                # Nome
                cell_nome = ws.cell(row=1, column=col, value=pessoa['Nome'])
                cell_nome.font = bold_font
                cell_nome.fill = header_fill
                cell_nome.alignment = center_align
                cell_nome.border = thin_border

                # Atividades
                cell_ativ = ws.cell(row=2, column=col,
                                    value=pessoa['Qtd Atividades'])
                cell_ativ.alignment = center_align
                cell_ativ.border = thin_border

                # Atos (se fornecido)
                if 'Qtd Atos' in pessoa:
                    cell_atos = ws.cell(row=3, column=col,
                                        value=pessoa['Qtd Atos'])
                    cell_atos.alignment = center_align
                    cell_atos.border = thin_border

            # Ajuste automático da largura das colunas
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max_len + 2

        wb.save(caminho_saida)

    except Exception as e:
        print(f"Erro ao atualizar a planilha: {e}")
